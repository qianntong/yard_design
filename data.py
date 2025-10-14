import pandas as pd
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def load_departure_data(departure_file):
    """
    Load departure schedule from Excel file
    Returns: DataFrame with Train, Scheduled Departure, and Bocks
    """
    print(f"Loading departure data from {departure_file}...")
    df = pd.read_excel(departure_file, sheet_name='Worksheet1')

    # Print all column names to debug
    print(f"Available columns: {df.columns.tolist()}")

    # Select relevant columns - note it's "Bocks" not "Blocks"
    df = df[['Train', 'Scheduled Departure', 'Bocks']].copy()

    # Remove rows with missing train names
    df = df.dropna(subset=['Train'])

    print(f"Loaded {len(df)} trains from departure schedule")
    return df


def load_hourly_counts(count_file):
    """
    Load hourly CAR ARRIVING counts from Excel file with multiple tabs
    Each tab represents a train, with columns: Train, Time, CAR_ARRIVING
    Returns: Dictionary with train names as keys and hourly counts as values
    """
    print(f"\nLoading hourly counts from {count_file}...")

    # Read all sheets from Excel file
    excel_file = pd.ExcelFile(count_file)
    sheet_names = excel_file.sheet_names
    print(f"Found {len(sheet_names)} sheets (trains): {sheet_names[:5]}..." if len(
        sheet_names) > 5 else f"Found {len(sheet_names)} sheets (trains): {sheet_names}")

    hourly_data = {}

    # Process each sheet (each sheet is a train)
    for sheet_name in sheet_names:
        train_name = sheet_name.strip()

        # Read the sheet
        df = pd.read_excel(count_file, sheet_name=sheet_name)

        # Debug: print structure of first sheet
        if sheet_name == sheet_names[0]:
            print(f"\nSample data structure from sheet '{sheet_name}':")
            print(f"Columns: {df.columns.tolist()}")
            print(f"Shape: {df.shape}")
            print(df.head())

        # Create hourly counts dictionary for this train
        hourly_counts = {}

        # Iterate through rows to get hourly data
        for idx, row in df.iterrows():
            time_str = str(row['Time'])  # e.g., "0:00", "1:00", etc.

            # Extract hour from time string
            try:
                hour = int(time_str.split(':')[0])
                car_arriving = row['CAR_ARRIVING']
                hourly_counts[hour] = int(car_arriving) if pd.notna(car_arriving) else 0
            except:
                print(f"Warning: Could not parse time '{time_str}' in sheet '{sheet_name}'")
                continue

        # Store the hourly data for this train
        hourly_data[train_name] = hourly_counts

        total_cars = sum(hourly_counts.values())
        print(f"Loaded train '{train_name}': {len(hourly_counts)} hours, {total_cars} total cars")

    print(f"\nLoaded hourly counts for {len(hourly_data)} trains")
    return hourly_data


def create_train_dataframe(train_name, hourly_counts, departure_time, bocks):
    """
    Create complete dataframe for a single train
    """
    hours = list(range(24))
    data = {
        'Train': [train_name] * 24,
        'Time': [f"{h}:00" for h in hours],
        'CAR_ARRIVING': [hourly_counts.get(h, 0) for h in hours]
    }

    df = pd.DataFrame(data)

    # Calculate DWELL_HOURS (hours remaining until end of day)
    df['DWELL_HOURS'] = df['Time'].apply(lambda x: 24 - int(x.split(':')[0]))

    # Calculate CAR_ARRIVING_X_DWELL
    df['CAR_ARRIVING_X_DWELL'] = df['CAR_ARRIVING'] * df['DWELL_HOURS']

    # Calculate totals
    total_car = df['CAR_ARRIVING'].sum()
    total_car_hours = df['CAR_ARRIVING_X_DWELL'].sum()

    # Initialize CAR_HOURS column
    df['CAR_HOURS'] = 0.0

    # Calculate CAR_HOURS backwards from hour 23
    for i in range(len(df) - 1, -1, -1):
        hour = int(df.loc[i, 'Time'].split(':')[0])
        car_arriving = df.loc[i, 'CAR_ARRIVING']

        if hour == 23:
            # Hour 23: total_car_hours - total_car + 24 * car_arriving(23)
            df.loc[i, 'CAR_HOURS'] = total_car_hours - total_car + 24 * car_arriving
        else:
            # Other hours: next hour's car_hours - total_car + 24 * car_arriving
            next_hour_idx = i + 1
            next_car_hours = df.loc[next_hour_idx, 'CAR_HOURS']
            df.loc[i, 'CAR_HOURS'] = next_car_hours - total_car + 24 * car_arriving

    # Add summary information
    df['TOTAL_CAR'] = total_car
    df['TOTAL_CAR_HOURS'] = total_car_hours
    df['DEPARTURE_TIME'] = departure_time
    df['BOCKS'] = bocks

    return df


def main(departure_file, count_file, output_file):
    """
    Main processing function
    """
    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Load departure schedule
    departure_df = load_departure_data(departure_file)

    # Load hourly counts from all tabs
    hourly_data = load_hourly_counts(count_file)

    print(f"\n{'=' * 60}")
    print(f"Processing {len(departure_df)} trains...")
    print(f"{'=' * 60}")

    # Create Excel writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    summary_data = []

    processed_count = 0
    skipped_count = 0

    # Process each train from departure schedule
    for idx, row in departure_df.iterrows():
        train_name = str(row['Train']).strip()
        departure_time = row['Scheduled Departure']
        bocks = row['Bocks']

        print(f"\n[{idx + 1}/{len(departure_df)}] Processing train: {train_name}")
        print(f"  Departure: {departure_time}, Bocks: {bocks}")

        # Get hourly counts for this train
        if train_name not in hourly_data:
            print(f"  Warning: No hourly data found for '{train_name}', skipping...")
            skipped_count += 1
            continue

        hourly_counts = hourly_data[train_name]

        # Create train dataframe
        train_df = create_train_dataframe(train_name, hourly_counts, departure_time, bocks)

        # Extract summary metrics
        total_car = train_df['TOTAL_CAR'].iloc[0]
        total_car_hours = train_df['TOTAL_CAR_HOURS'].iloc[0]
        avg_car_hours = total_car_hours / total_car if total_car > 0 else 0

        min_car_hours_idx = train_df['CAR_HOURS'].idxmin()
        min_car_hours = train_df.loc[min_car_hours_idx, 'CAR_HOURS']
        min_car_hours_time = train_df.loc[min_car_hours_idx, 'Time']

        summary_data.append({
            'Train': train_name,
            'DEPT_TIME': departure_time,
            'Bocks': bocks,
            'TOTAL_CAR': total_car,
            'TOTAL_CAR_HOURS': total_car_hours,
            'AVG_CAR_HOURS': round(avg_car_hours, 2),
            'MIN_CAR_HOURS': round(min_car_hours, 2),
            'MIN_CAR_HOURS_TIME': min_car_hours_time
        })

        # Create safe sheet name (Excel has 31 character limit)
        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', train_name)[:31]
        train_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

        processed_count += 1
        print(
            f"  Completed! Total Cars: {total_car}, Total Car Hours: {total_car_hours:.2f}, Avg: {avg_car_hours:.2f}")

    writer.close()

    # Create summary sheet
    if summary_data:
        summary_df = pd.DataFrame(summary_data)
        wb = openpyxl.load_workbook(output_file)

        if 'Summary' in wb.sheetnames:
            del wb['Summary']
        ws_summary = wb.create_sheet('Summary', 0)

        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_file)
        print(f"\n{'=' * 60}")
        print(f" Summary sheet created with {len(summary_data)} trains")
        print(f"{'=' * 60}")

    print(f"\n Processing Summary:")
    print(f"  - Total trains in departure schedule: {len(departure_df)}")
    print(f"  - Successfully processed: {processed_count}")
    print(f"  - Skipped (no data): {skipped_count}")
    print(f"\n Results saved to: {output_file}")


if __name__ == "__main__":
    departure_file = "data/TH-Outbound-Train-Plan-2025.xlsx"
    count_file = "data/alt_2_blocks_sum.xlsx"
    output_file = "results/yard_chart_results_alt_2.xlsx"

    main(departure_file, count_file, output_file)