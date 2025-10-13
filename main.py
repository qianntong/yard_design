import pandas as pd
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def parse_spare_blocks(spare_str):
    """
    Parse block information from SPARE columns
    Example: "2 CHBR 1 CHG" -> {'CHBR': 2, 'CHG': 1}
    """
    if pd.isna(spare_str) or spare_str == '':
        return {}

    blocks = {}
    parts = str(spare_str).split()
    i = 0
    while i < len(parts):
        if parts[i].isdigit():
            count = int(parts[i])
            if i + 1 < len(parts):
                block_name = parts[i + 1]
                blocks[block_name] = blocks.get(block_name, 0) + count
                i += 2
            else:
                i += 1
        else:
            i += 1

    return blocks


def parse_time_from_column(time_str):
    """
    Parse time column and return hour (0-23)
    Example: "0:00-0:15" -> 0
    """
    if pd.isna(time_str):
        return None

    time_str = str(time_str).strip()
    match = re.match(r'(\d+):(\d+)', time_str)
    if match:
        hour = int(match.group(1))
        return hour
    return None


def get_blocks_from_departure(blocks_str):
    """
    Parse blocks from departure table
    Example: "RLK, ESTR" -> ['RLK', 'ESTR']
    """
    if pd.isna(blocks_str):
        return []

    blocks = [b.strip() for b in str(blocks_str).split(',')]
    return blocks


def find_earliest_pull_time(yard_plan, train_name):
    """
    Find the earliest pull time for a specified train in yard_plan
    """
    pull_columns = [col for col in yard_plan.columns if col.startswith('Pull')]

    earliest_time = None
    earliest_hour = None

    for _, row in yard_plan.iterrows():
        for col in pull_columns:
            cell_value = row[col]
            # Fix: Convert to string and handle NaN values
            if pd.isna(cell_value):
                continue
            cell_value = str(cell_value)

            if train_name in cell_value:
                hour = parse_time_from_column(row.get('Time', ''))
                if hour is not None:
                    # Handle midnight crossing case
                    if earliest_hour is None:
                        earliest_hour = hour
                        earliest_time = row['Time']
                    elif hour == 23 and earliest_hour < 2:
                        # 23:xx is earlier than 0:xx-1:xx (crossing midnight)
                        earliest_hour = hour
                        earliest_time = row['Time']
                    elif hour < earliest_hour and not (earliest_hour == 23 and hour < 2):
                        earliest_hour = hour
                        earliest_time = row['Time']

    return earliest_time, earliest_hour


def calculate_car_arriving(yard_plan, blocks, earliest_hour):
    """
    Calculate CAR ARRIVING for each hour
    Exclude the row at earliest_hour (clearing operation)
    """
    hourly_counts = {h: 0 for h in range(24)}

    # Find all relevant block columns and spare columns
    block_columns = []
    spare_columns = []

    for col in yard_plan.columns:
        if col.startswith('SPARE'):
            spare_columns.append(col)
        elif col in blocks:
            block_columns.append(col)

    # Iterate through each row, excluding the row at earliest_hour
    for _, row in yard_plan.iterrows():
        hour = parse_time_from_column(row.get('Time', ''))

        if hour is None or hour == earliest_hour:
            continue

        count = 0

        # Count direct block columns
        for block_col in block_columns:
            val = row.get(block_col, 0)
            if pd.notna(val) and str(val).replace('.', '').replace('-', '').isdigit():
                count += int(float(val))
                # print(f"column count: {count}")

        # Count target blocks in spare columns
        for spare_col in spare_columns:
            spare_value = row.get(spare_col, '')
            spare_blocks = parse_spare_blocks(spare_value)
            for block in blocks:
                if block in spare_blocks:
                    count += spare_blocks[block]


        hourly_counts[hour] += max(0,count)


    print(f"count: {count}; hourly counts: {hourly_counts}")
    return hourly_counts


def create_train_dataframe(train_name, hourly_counts, departure_time):
    """
    Create complete dataframe for a single train
    """
    hours = list(range(24))
    data = {
        'Train': [train_name] * 24,
        'Time': [f"{h}:00" for h in hours],
        'CAR_ARRIVING': [hourly_counts[h] for h in hours]
    }

    print(f"data for {train_name}: {data}")
    df = pd.DataFrame(data)

    df['DWELL_HOURS'] = df['Time'].apply(lambda x: 24 - int(x.split(':')[0]))
    df['CAR_ARRIVING_X_DWELL'] = df['CAR_ARRIVING'] * df['DWELL_HOURS']
    total_car = df['CAR_ARRIVING'].sum()
    total_car_hours = df['CAR_ARRIVING_X_DWELL'].sum()
    df['CAR_HOURS'] = 0.0


    for i in range(len(df) - 1, -1, -1): # Work backwards from hour 23
        hour = int(df.loc[i, 'Time'].split(':')[0])
        car_arriving = df.loc[i, 'CAR_ARRIVING']

        if hour == 23:
            # 23pm: total_car_hours - total_car + 24 * car_arriving(23pm)
            df.loc[i, 'CAR_HOURS'] = total_car_hours - total_car + 24 * car_arriving
        else:
            # Other hours: previous hour's car_hours - total_car + current dwell * car_arriving
            next_hour_idx = i + 1
            prev_car_hours = df.loc[next_hour_idx, 'CAR_HOURS']
            df.loc[i, 'CAR_HOURS'] = prev_car_hours - total_car + 24 * car_arriving

    # Add summary information
    df['TOTAL_CAR'] = total_car
    df['TOTAL_CAR_HOURS'] = total_car_hours
    df['DEPARTURE_TIME'] = departure_time

    return df


def main(departure_file, yard_plan_file, output_file):
    print("Reading departure table...")
    departure_df = pd.read_excel(departure_file, sheet_name='Worksheet1')

    print("Reading yard_plan...")
    yard_plan = pd.read_csv(yard_plan_file)

    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created directory: {output_dir}")

    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    summary_data = []

    for idx, row in departure_df.iterrows():
        train_name = str(row['Train'])
        departure_time = row['Scheduled Departure']
        blocks_str = row['Bocks']

        print(f"\nProcessing train: {train_name}")
        print(f"  Scheduled departure: {departure_time}")
        print(f"  Blocks: {blocks_str}")

        blocks = get_blocks_from_departure(blocks_str)

        if not blocks:
            print(f"  Warning: {train_name} has no valid blocks information")
            continue

        earliest_time, earliest_hour = find_earliest_pull_time(yard_plan, train_name)

        if earliest_hour is None:
            print(f"  Warning: {train_name} not found in yard_plan")
            continue

        print(f"  Earliest pull time: {earliest_time} (hour: {earliest_hour})")

        hourly_counts = calculate_car_arriving(yard_plan, blocks, earliest_hour)

        train_df = create_train_dataframe(train_name, hourly_counts, departure_time)

        # Extract summary metrics
        total_car = train_df['TOTAL_CAR'].iloc[0]
        total_car_hours = train_df['TOTAL_CAR_HOURS'].iloc[0]
        avg_car_hours = total_car_hours / total_car if total_car > 0 else 0

        min_car_hours_idx = train_df['CAR_HOURS'].idxmin()
        min_car_hours = train_df.loc[min_car_hours_idx, 'CAR_HOURS']
        min_car_hours_time = train_df.loc[min_car_hours_idx, 'Time']

        summary_data.append({
            'Train': train_name,
            'DEPT_TIME': departure_time,
            'TOTAL_CAR': total_car,
            'TOTAL_CAR_HOURS': total_car_hours,
            'AVG_CAR_HOURS': round(avg_car_hours,2),
            'MIN_CAR_HOURS': min_car_hours,
            'MIN_CAR_HOURS_TIME': min_car_hours_time
        })

        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', train_name)[:31]
        train_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

        print(f"  Completed! Total Cars: {total_car}, Total Car Hours: {total_car_hours:.2f}, Avg: {avg_car_hours:.2f}")

    writer.close()

    # Create summary sheet
    if summary_data:
        summary_df = pd.DataFrame(summary_data)
        wb = openpyxl.load_workbook(output_file)

        if 'Summary' in wb.sheetnames:
            del wb['Summary']
        ws_summary = wb.create_sheet('Summary', 0)

        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_file)
        print(f"\nSummary sheet created with {len(summary_data)} trains")

    print(f"\nResults saved to: {output_file}")

if __name__ == "__main__":
    departure_file = "data/TH-Outbound-Train-Plan-2025.xlsx"
    yard_plan_file = "data/alt_3.csv"
    output_file = "results/yard_chart_results_alt_3.xlsx"

    main(departure_file, yard_plan_file, output_file)